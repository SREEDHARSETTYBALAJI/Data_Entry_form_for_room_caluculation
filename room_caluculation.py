from tkinter import *
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime


displayed_names = []

def insertitem():
    if not os.path.exists('Book1.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Price", "Item", "Name"])
        wb.save('Book1.xlsx')
    wb = load_workbook('Book1.xlsx')
    ws = wb.active
    entry_date = date.get()
    entry_price = price.get()
    entry_item = item.get()
    entry_name = name.get()
    ws.append([entry_date, entry_price, entry_item, entry_name])
    wb.save('Book1.xlsx')

    date.delete(0, END)
    price.delete(0, END)
    item.delete(0, END)
    name.delete(0, END)
    print("Data added successfully!")

def datashown():
    global displayed_names
    displayed_names = []  
    display.delete(1.0, END)
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    if os.path.exists('Book1.xlsx'):
        wb = load_workbook('Book1.xlsx')
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            try:
                row_date = datetime.strptime(row[0], "%Y-%m-%d")  
                if row_date.month == current_month and row_date.year == current_year:
                    display.insert(END, f" {row[0]},  {row[1]},  {row[2]},  {row[3]}\n")
                    displayed_names.append(row[3])  
            except ValueError:
                continue  

def clear_display():
    display.delete(1.0, END)

def calculate_total_price():
    total_price = 0
    if os.path.exists('Book1.xlsx'):
        wb = load_workbook('Book1.xlsx')
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):  
            try:
                price_value = float(row[1])  
                name_value = row[3]
                if name_value in displayed_names:  
                    total_price += price_value
            except (ValueError, TypeError):
                continue  
    display.delete(1.0, END)
    display.insert(END, f"Total Price for displayed names: {total_price}\n")

def show_history():
    display.delete(1.0, END)
    if os.path.exists('Book1.xlsx'):
        wb = load_workbook('Book1.xlsx')
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            display.insert(END, f" {row[0]},  {row[1]},  {row[2]},  {row[3]}\n")

home = Tk()
home.title("Data Entry Form")

date = Entry(home)
date.grid(row=0, column=1)
price = Entry(home)
price.grid(row=1, column=1)
item = Entry(home)
item.grid(row=2, column=1)
name = Entry(home)
name.grid(row=3, column=1)

Label(home, text="Date (YYYY-MM-DD)").grid(row=0, column=0)
Label(home, text="Price").grid(row=1, column=0)
Label(home, text="Item").grid(row=2, column=0)
Label(home, text="Name").grid(row=3, column=0)

enter = Button(home, text="Enter", command=insertitem)
enter.grid(row=4, column=0)

show = Button(home, text="Show Data", command=datashown)
show.grid(row=4, column=1)

clear = Button(home, text="Clear", command=clear_display)
clear.grid(row=4, column=2)

total_price_button = Button(home, text="Total Price", command=calculate_total_price)
total_price_button.grid(row=4, column=3)

history_button = Button(home, text="History", command=show_history)
history_button.grid(row=4, column=4)

display = Text(home, height=10, width=70)

display.grid(row=5, column=0, columnspan=5)
home.mainloop()
